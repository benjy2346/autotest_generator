
from openpyxl import Workbook,load_workbook  
import os
from qianfan import  ChatCompletion
from datetime import datetime
from config import api_key, secrete_key,output_path

cache_conversation = []
comp = ChatCompletion(ak = api_key,sk= secrete_key,model='ERNIE-Speed-8K')
time_stamp = datetime.now().strftime('%Y-%m-%d_%H-%M')
pth = output_path + f'{time_stamp}.xlsx'



def wenxin_engine(prompt,iter, accumulate):
    if accumulate == False:
        cache_conversation=[]
        
    results=[]
    results.append( engine_init(prompt))
    
    for _ in range(iter):
        results.append(generate_more())
    write_to_excel(results, accumulate, path=pth)
    
    # 尝试打开
    try:
        os.startfile(os.path.abspath(pth))
    except:
        print('请关闭excel文档')
    
def generate_more():
    msg  = {
            'role':'user',
            "content": '提供更多测试用例,应该包含以下信息：测试模块、测试标题、前置条件、测试步骤、预期结果、实际结果、测试方法等。 测试用例输出的格式使用之前的表格格式。'
        }
    cache_conversation.append(msg)
    resp = comp.do(
        messages=cache_conversation
    )
    result = resp['body']['result']
    cache_conversation.append({
            'role':'assistant',
            'content':result})
    return result[result.find('|'):result.rfind('|')]

def engine_init(prompt):
    msg = {
            'role':'user',
            "content": prompt
        }
    cache_conversation.append(msg)
    resp = comp.do(
        messages=cache_conversation
    )
    
    init_result = resp['body']['result']
    print(init_result)
    cache_conversation.append({
            'role':'assistant',
            'content':init_result})
    return init_result[init_result.find('|'):init_result.rfind('|')]

    
def find_first_row_with_data(wb):  
    # 加载工作簿  
    wb = load_workbook(pth)
    workbook = wb 
    # 选择第一个工作表  
    sheet = workbook.active  
  
      
    # 如果没有找到数据，返回None或你想要的默认值  
    return sheet.max_row 
def write_to_excel( results, accumulate,path):
    

    # 从第一行第一列开始写入  
    row = 1  
    col = 0
    if accumulate and os.path.exists(pth):
            wb = load_workbook(pth)
            row = find_first_row_with_data(wb)+2
    else:
            wb = Workbook()
    ws = wb.active 

    for r in results:
        
        for words in r.split('|'): 
            if words == '': 
                print('跳过空白')
                continue
            ws.cell(row=row, column=col+1, value=words) 
            
            col = (col +1)
            
            if words == '\n':
                row += 1  # 移动到下一行  
                col = 0
        # 空一行方便看
        row+= 1
        col=0
        
    os.makedirs(output_path,exist_ok=True)
    wb.save(path)
    

        
        
    