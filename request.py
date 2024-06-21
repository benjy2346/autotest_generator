import requests
import qianfan
from openpyxl import Workbook  
from openpyxl.utils import get_column_letter
import os
import qianfan
from qianfan import ChatCompletion, Completion
from qianfan.common import Prompt
import time
from config import api_key, secrete_key,output_path
USE_CACHE= False



# 接下来就可以调用 SDK 的所有功能
def get_access_key(api_key, secrete_key):
    data = {"client_id":api_key,
            "client_secret": secrete_key,
            "grant_type":"client_credentials"}
    response = requests.post('https://aip.baidubce.com/oauth/2.0/token',data).json()
    return response['access_token']

def connect2llm(insight, info_blocks,personality_requirements,
                min_data_num,testing_modules_included,testing_methods,capacity):
    print({
        'statement':"请根据该需求生成测试用例。" +f'确保你的测试用例总数至少为{min_data_num}条。',
        'capacity':capacity,
        'insight':insight,
        'personality':f"""你的回答应该细致、具体，并针对每个可能出现的边界情况编写尽可能多并且全面的测试用例。"""+personality_requirements,
        'experiment':f'要求测试用例包含以下信息：{info_blocks}。测试用例输出的格式使用表格格式。'+testing_modules_included + testing_methods + "请确保没有空格。"
    })
    comp = Completion(ak = api_key,sk= secrete_key,model='ERNIE-Speed-8K')

    crispe_prompt = Prompt(Prompt.crispe_prompt(
        statement="请根据该需求生成测试用例。" +f'确保你的测试用例总数至少为{min_data_num}条。',
        capacity=capacity,
        insight=insight,
        personality=f"""你的回答应该细致、具体，并针对每个可能出现的边界情况编写尽可能多并且全面的测试用例。"""+personality_requirements,
        experiment=f'要求测试用例包含以下信息：、{info_blocks}。测试用例输出的格式使用表格格式。'+testing_modules_included + testing_methods + "请确保没有空格。"
        ),
        identifier="{{}}"
    )
    resp = comp.do(
        prompt=crispe_prompt.render()[0]
    )
    result = resp['body']['result']
    result =  result[result.find('|'):result.rfind('|')]
    
    
    wb = Workbook()  
    ws = wb.active 
  
    # 假设我们想要从第一行第一列开始写入  
    row = 1  
    col = 0

    # cache
    # if not os.path.exists('cache.txt'):
    #     file = open('cache.txt','w+',encoding='utf-8')
    #     file.write(result)    
    #     file.close()
        
    # elif open('cache.txt','r',encoding='utf-8').read() == '':
    #     file = open('cache.txt','w+',encoding='utf-8')
    #     file.write(result) 
    #     file.close()
    # else:
    #     with open('cache.txt', 'r', encoding='utf-8') as f:  
    #         result = f.read()

    # 遍历文字列表，并将它们写入工作表  
    for words in result.split('|'): 
        if words == '': 
            print('跳过空白')
            continue
        ws.cell(row=row, column=col+1, value=words) 
        
        col = (col +1)
        
        if words == '\n':
            row += 1  # 移动到下一行  
            col = 0
            
    # 保存工作簿到文件  
    if not os.path.exists(output_path):  
        os.makedirs(output_path)
    time_stamp = int( time.time())
    pth = output_path + f'{time_stamp}.xlsx'
    wb.save(pth)
    
    try:
        
        os.startfile(os.path.abspath(pth))
    except:
        print('请关闭excel文档')